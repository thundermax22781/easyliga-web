from fastapi import FastAPI, APIRouter, HTTPException, Query, UploadFile, File, Form
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import random
import io
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, date, timezone
import openpyxl

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# --- Models ---

VALID_ROLES = ["Portiere", "Difensore", "Centrocampista", "Attaccante"]
ROLE_ORDER = {"Portiere": 0, "Difensore": 1, "Centrocampista": 2, "Attaccante": 3}

VALID_COLORS = ["Bianca", "Rossa", "Gialla", "Nera", "Verde"]
VALID_MATCH_TYPES = [5, 6, 7, 8, 9, 10, 11]

# --- Group Models ---

class GroupCreate(BaseModel):
    name: str

class GroupUpdate(BaseModel):
    name: str

class GroupResponse(BaseModel):
    id: str
    name: str
    player_count: int = 0
    created_at: str

class PlayerCreate(BaseModel):
    name: str
    surname: str
    nickname: str
    date_of_birth: str  # ISO format YYYY-MM-DD
    photo: Optional[str] = None  # base64 string
    role: str
    strength: float = Field(ge=1, le=10)
    group_id: str

class PlayerUpdate(BaseModel):
    name: Optional[str] = None
    surname: Optional[str] = None
    nickname: Optional[str] = None
    date_of_birth: Optional[str] = None
    photo: Optional[str] = None
    role: Optional[str] = None
    strength: Optional[float] = Field(default=None, ge=1, le=10)

class PlayerResponse(BaseModel):
    id: str
    name: str
    surname: str
    nickname: str
    date_of_birth: str
    age: int
    photo: Optional[str] = None
    role: str
    strength: float
    created_at: str
    updated_at: str

class TeamGenerateRequest(BaseModel):
    player_ids: List[str]
    players_per_team: int = 5
    team_a_name: str = "Squadra A"
    team_b_name: str = "Squadra B"
    team_a_color: str = "Bianca"
    team_b_color: str = "Rossa"

class TeamResponse(BaseModel):
    team_a: List[PlayerResponse]
    team_b: List[PlayerResponse]
    team_a_total_strength: float
    team_b_total_strength: float
    team_a_avg_age: float
    team_b_avg_age: float
    team_a_name: str
    team_b_name: str
    team_a_color: str
    team_b_color: str

def calculate_age(dob_str: str) -> int:
    dob = date.fromisoformat(dob_str)
    today = date.today()
    age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
    return age

def player_doc_to_response(doc: dict) -> dict:
    return {
        "id": doc["id"],
        "name": doc["name"],
        "surname": doc["surname"],
        "nickname": doc["nickname"],
        "date_of_birth": doc["date_of_birth"],
        "age": calculate_age(doc["date_of_birth"]),
        "photo": doc.get("photo"),
        "role": doc["role"],
        "strength": doc["strength"],
        "created_at": doc["created_at"],
        "updated_at": doc["updated_at"],
    }

# --- Routes ---

@api_router.get("/")
async def root():
    return {"message": "Calcetto Manager API"}

# --- Group Routes ---

@api_router.post("/groups", response_model=GroupResponse)
async def create_group(group: GroupCreate):
    if not group.name.strip():
        raise HTTPException(status_code=400, detail="Il nome del gruppo non può essere vuoto")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "name": group.name.strip(),
        "created_at": now,
    }
    await db.groups.insert_one(doc)
    return {"id": doc["id"], "name": doc["name"], "player_count": 0, "created_at": doc["created_at"]}

@api_router.get("/groups", response_model=List[GroupResponse])
async def get_groups():
    groups = await db.groups.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    result = []
    for g in groups:
        count = await db.players.count_documents({"group_id": g["id"]})
        result.append({"id": g["id"], "name": g["name"], "player_count": count, "created_at": g["created_at"]})
    return result

@api_router.put("/groups/{group_id}", response_model=GroupResponse)
async def update_group(group_id: str, group: GroupUpdate):
    doc = await db.groups.find_one({"id": group_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Gruppo non trovato")
    await db.groups.update_one({"id": group_id}, {"$set": {"name": group.name.strip()}})
    updated = await db.groups.find_one({"id": group_id}, {"_id": 0})
    count = await db.players.count_documents({"group_id": group_id})
    return {"id": updated["id"], "name": updated["name"], "player_count": count, "created_at": updated["created_at"]}

@api_router.delete("/groups/{group_id}")
async def delete_group(group_id: str):
    result = await db.groups.delete_one({"id": group_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Gruppo non trovato")
    await db.players.delete_many({"group_id": group_id})
    return {"message": "Gruppo e giocatori eliminati"}

# --- Player Routes ---

@api_router.post("/players", response_model=PlayerResponse)
async def create_player(player: PlayerCreate):
    if player.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Ruolo non valido. Scegli tra: {VALID_ROLES}")
    if player.strength * 2 != int(player.strength * 2):
        raise HTTPException(status_code=400, detail="La forza deve essere un multiplo di 0.5")
    
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "name": player.name,
        "surname": player.surname,
        "nickname": player.nickname,
        "date_of_birth": player.date_of_birth,
        "photo": player.photo,
        "role": player.role,
        "strength": player.strength,
        "group_id": player.group_id,
        "created_at": now,
        "updated_at": now,
    }
    await db.players.insert_one(doc)
    return player_doc_to_response(doc)

@api_router.get("/players", response_model=List[PlayerResponse])
async def get_players(
    search: Optional[str] = Query(None),
    role: Optional[str] = Query(None),
    group_id: Optional[str] = Query(None),
    min_strength: Optional[float] = Query(None, ge=1, le=10),
    max_strength: Optional[float] = Query(None, ge=1, le=10),
    sort_by: Optional[str] = Query("nickname"),
    sort_order: Optional[str] = Query("asc"),
):
    query = {}
    if group_id:
        query["group_id"] = group_id
    if search:
        query["$or"] = [
            {"nickname": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}},
            {"surname": {"$regex": search, "$options": "i"}},
        ]
    if role and role in VALID_ROLES:
        query["role"] = role
    if min_strength is not None or max_strength is not None:
        strength_q = {}
        if min_strength is not None:
            strength_q["$gte"] = min_strength
        if max_strength is not None:
            strength_q["$lte"] = max_strength
        query["strength"] = strength_q

    sort_dir = 1 if sort_order == "asc" else -1
    valid_sorts = ["nickname", "strength", "role", "created_at"]
    sort_field = sort_by if sort_by in valid_sorts else "nickname"

    docs = await db.players.find(query, {"_id": 0}).sort(sort_field, sort_dir).to_list(1000)
    return [player_doc_to_response(d) for d in docs]

@api_router.get("/players/template")
async def download_template():
    """Download Excel template for player import"""
    from starlette.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Giocatori"
    headers = ["Nickname", "Nome", "Cognome", "Data di Nascita (AAAA-MM-GG)", "Ruolo", "Forza (1-10)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    ws.cell(row=2, column=1, value="SuperMario")
    ws.cell(row=2, column=2, value="Mario")
    ws.cell(row=2, column=3, value="Rossi")
    ws.cell(row=2, column=4, value="1995-03-15")
    ws.cell(row=2, column=5, value="Attaccante")
    ws.cell(row=2, column=6, value=7.5)
    
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_giocatori.xlsx"}
    )

@api_router.post("/players/import")
async def import_players(file: UploadFile = File(...), group_id: str = Form(...)):
    """Import players from Excel file"""
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Il file deve essere in formato Excel (.xlsx)")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Impossibile leggere il file Excel")
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Il file Excel è vuoto (nessuna riga dopo l'intestazione)")
    
    created = []
    errors = []
    now = datetime.now(timezone.utc).isoformat()
    
    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        try:
            nickname = str(row[0] or '').strip()
            name = str(row[1] or '').strip()
            surname = str(row[2] or '').strip()
            dob_raw = row[3]
            role = str(row[4] or '').strip()
            strength_raw = row[5]
            
            if not nickname:
                errors.append(f"Riga {i}: Nickname mancante")
                continue
            if not name:
                name = nickname
            if not surname:
                surname = ""
            
            dob_str = ""
            if isinstance(dob_raw, datetime):
                dob_str = dob_raw.strftime("%Y-%m-%d")
            elif isinstance(dob_raw, date):
                dob_str = dob_raw.isoformat()
            elif isinstance(dob_raw, str):
                dob_str = dob_raw.strip()
            
            if not dob_str:
                errors.append(f"Riga {i} ({nickname}): Data di nascita mancante")
                continue
            
            try:
                date.fromisoformat(dob_str)
            except ValueError:
                errors.append(f"Riga {i} ({nickname}): Data non valida '{dob_str}'")
                continue
            
            role_map = {
                "portiere": "Portiere", "por": "Portiere", "p": "Portiere",
                "difensore": "Difensore", "dif": "Difensore", "d": "Difensore",
                "centrocampista": "Centrocampista", "cen": "Centrocampista", "c": "Centrocampista",
                "attaccante": "Attaccante", "att": "Attaccante", "a": "Attaccante",
            }
            role_normalized = role_map.get(role.lower(), role)
            if role_normalized not in VALID_ROLES:
                errors.append(f"Riga {i} ({nickname}): Ruolo '{role}' non valido")
                continue
            
            try:
                strength = float(strength_raw)
                if strength < 1 or strength > 10:
                    errors.append(f"Riga {i} ({nickname}): Forza deve essere tra 1 e 10")
                    continue
                strength = round(strength * 2) / 2
            except (TypeError, ValueError):
                errors.append(f"Riga {i} ({nickname}): Forza non valida '{strength_raw}'")
                continue
            
            doc = {
                "id": str(uuid.uuid4()),
                "name": name,
                "surname": surname,
                "nickname": nickname,
                "date_of_birth": dob_str,
                "photo": None,
                "role": role_normalized,
                "strength": strength,
                "group_id": group_id,
                "created_at": now,
                "updated_at": now,
            }
            await db.players.insert_one(doc)
            created.append(player_doc_to_response(doc))
        except Exception as e:
            errors.append(f"Riga {i}: Errore imprevisto - {str(e)}")
    
    return {
        "imported": len(created),
        "errors": errors,
        "players": created,
    }

@api_router.get("/players/{player_id}", response_model=PlayerResponse)
async def get_player(player_id: str):
    doc = await db.players.find_one({"id": player_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Giocatore non trovato")
    return player_doc_to_response(doc)

@api_router.put("/players/{player_id}", response_model=PlayerResponse)
async def update_player(player_id: str, player: PlayerUpdate):
    doc = await db.players.find_one({"id": player_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Giocatore non trovato")
    
    update_data = {k: v for k, v in player.dict().items() if v is not None}
    if "role" in update_data and update_data["role"] not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Ruolo non valido. Scegli tra: {VALID_ROLES}")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.players.update_one({"id": player_id}, {"$set": update_data})
    
    updated = await db.players.find_one({"id": player_id}, {"_id": 0})
    return player_doc_to_response(updated)

@api_router.delete("/players/{player_id}")
async def delete_player(player_id: str):
    result = await db.players.delete_one({"id": player_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Giocatore non trovato")
    return {"message": "Giocatore eliminato"}

@api_router.post("/generate-teams", response_model=TeamResponse)
async def generate_teams(req: TeamGenerateRequest):
    if len(req.player_ids) < 2:
        raise HTTPException(status_code=400, detail="Servono almeno 2 giocatori")
    
    players = []
    for pid in req.player_ids:
        doc = await db.players.find_one({"id": pid}, {"_id": 0})
        if doc:
            players.append(doc)
    
    if len(players) < 2:
        raise HTTPException(status_code=400, detail="Giocatori non trovati")
    
    # Balanced team generation: sort by strength, alternate assignment
    sorted_players = sorted(players, key=lambda p: p["strength"], reverse=True)
    
    team_a = []
    team_b = []
    sum_a = 0
    sum_b = 0
    
    for p in sorted_players:
        if sum_a <= sum_b:
            team_a.append(p)
            sum_a += p["strength"]
        else:
            team_b.append(p)
            sum_b += p["strength"]
    
    # Sort within teams by role order: Portiere, Difensore, Centrocampista, Attaccante
    team_a.sort(key=lambda p: ROLE_ORDER.get(p["role"], 99))
    team_b.sort(key=lambda p: ROLE_ORDER.get(p["role"], 99))
    
    avg_a = sum_a / len(team_a) if team_a else 0
    avg_b = sum_b / len(team_b) if team_b else 0
    
    age_a = sum(calculate_age(p["date_of_birth"]) for p in team_a) / len(team_a) if team_a else 0
    age_b = sum(calculate_age(p["date_of_birth"]) for p in team_b) / len(team_b) if team_b else 0
    
    return {
        "team_a": [player_doc_to_response(p) for p in team_a],
        "team_b": [player_doc_to_response(p) for p in team_b],
        "team_a_total_strength": round(sum_a, 1),
        "team_b_total_strength": round(sum_b, 1),
        "team_a_avg_age": round(age_a, 1),
        "team_b_avg_age": round(age_b, 1),
        "team_a_name": req.team_a_name,
        "team_b_name": req.team_b_name,
        "team_a_color": req.team_a_color,
        "team_b_color": req.team_b_color,
    }

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
